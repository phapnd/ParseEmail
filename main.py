import imapclient
import pyzmail
import pprint
import os
import openpyxl
from pyzmail import PyzMessage
import email
from time import sleep
from imapclient import SEEN


def connect(server, email, password):
	"""
	Get Imap account cconnection with server
	"""
	imapObj = imapclient.IMAPClient(server, use_uid= True, ssl=True)
	imapObj.login(email, password)
	#imapObj.select_folder('INBOX', readonly=True)
	imapObj.select_folder('INBOX')
	return imapObj

def print_tree(account):
    """
    Print folder tree
    """
    print(account.list_folders())

def get_unread(account):
    """
    Get unread emails
    """
    UIDs = account.search(['UNSEEN'])
    return UIDs

def get_pyzmail(account,id_email):
    """
    Get pyzmail email
    """
    raw_msg = account.fetch([id_email], [b'BODY[]', 'FLAGS'])
    msg =  pyzmail.PyzMessage.factory(raw_msg[id_email][b'BODY[]'])
    return msg

def get_stmail(account,id_email):
	"""
	Get standard format email
	"""
	messageParts = account.fetch([id_email], [b'BODY[]', 'RFC822'])
	#emailBody = messageParts[0][1]
	#print("\nGET_STMAIL")
	#print(messageParts)
	mail = email.message_from_bytes(messageParts[id_email][b'RFC822'])
	return mail

def set_read_email(account,id_email):
	"""
	Set email as read - seen
	"""
	#account.store(data[0].replace(' ',','),'+FLAGS','\Seen
	account.set_flags(id_email,[SEEN])

def download_attachment(mail,folder_name):
	"""
	Download attachment included in email, save in folder_name
	"""
	global fileName
	global pathAttachment
	for part in mail.walk():
		if part.get_content_maintype() == 'multipart':
			# print part.as_string()
			continue
		if part.get('Content-Disposition') is None:
			# print part.as_string()
			continue
		fileName = part.get_filename()
		if bool(fileName):
		    pathAttachment = os.path.join(folder_name, 'attachments')
		    filePath = os.path.join(folder_name, 'attachments', fileName)
		    if not os.path.isfile(filePath) :
		        print(fileName)
		        #print(part.get_payload(decode=True))
		        fp = open(filePath, 'wb')
		        fp.write(part.get_payload(decode=True))
		        fp.close()

def save_excel(fileName, folderName, titleMail, fromName, fromEmail, fromDate, pathToAttachment):
	if bool(fileName):
		filePath = os.path.join(folderName, 'attachments', fileName)
		#workbook = xlsxwriter.Workbook(fileName)
		if not os.path.isfile(filePath):
			workbook = openpyxl.Workbook()
			sheet = workbook.active
			sheet["A1"] = "Title Mail"
			sheet["B1"] = "From Name"
			sheet["C1"] = "From Mail"
			sheet["D1"] = "From Date"
			sheet["E1"] = "Path to Attachments"
			new_row = [titleMail, fromName, fromEmail, fromDate, pathToAttachment]
			sheet.append(new_row)
			workbook.save(filename=filePath)
		else:
			workbook=openpyxl.load_workbook(filename = filePath)
			sheet = workbook.active
			new_row = [titleMail, fromName, fromEmail, fromDate, pathToAttachment]
			sheet.append(new_row)
			workbook.save(filename=filePath)


def main():
    global fileName
    global filePath
    # Connection details
    server = 'imap.gmail.com'
    email = input("Email: \n")
    password = input("Password: \n")

    account = connect(server, email, password)

    # Print all subject emails
    UIDs = get_unread(account)
    for uid in UIDs:
    	msg = get_pyzmail(account,uid)
    	#print("\nFrom: " + msg.get_address("from")[1])
    	#print("\nName: " + msg.get_address("from")[0])
    	#print("\nTo: " + ''.join(msg.get_address("to")))
    	#print("\nSubject: " + msg.get_subject())
    	fromEmail = msg.get_address("from")[1]
    	fromName = msg.get_address("from")[0]
    	titleMail = msg.get_subject()

    	body_msg_txt = lambda msg: "" if  msg.text_part == None else  msg.text_part.get_payload().decode(msg.text_part.charset)
    	print("\nBody: ")
    	print(body_msg_txt(msg))
    	body_msg_html = lambda msg: "" if msg.html_part == None else msg.html_part.get_payload().decode(msg.html_part.charset)
    	print("\n")
    	st_mail = get_stmail(account,uid)
    	download_attachment(st_mail,r"C://Users//hanht//Downloads//")
    	#print(body_msg_html(msg))
    	# Download Attachment
    	#print("\nDate: ")
    	#print(msg.get_decoded_header('date'))
    	fromDate = msg.get_decoded_header('date')
    	set_read_email(account,uid)
    	save_excel("Result.xlsx", r"C://Users//hanht//Downloads//", titleMail, fromName, fromEmail, fromDate, pathAttachment)
    	#print(email.utils.parsedate(msg.get_decoded_header('date')))
    	sleep(2)


    account.logout()

if __name__ == '__main__':
    main()
